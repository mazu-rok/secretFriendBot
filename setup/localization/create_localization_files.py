import openpyxl
import sys

workbook = openpyxl.load_workbook(sys.argv[1])
worksheet = workbook.worksheets[0]
out_files = []
for col in worksheet.iter_cols(min_col=2, max_col=worksheet.max_column):
    if col[0].value == "default":
        out_files.append(open("messages.properties", "a+"))
    else:
        out_files.append(open("messages_{}.properties".format(col[0].value), "a+", encoding="utf-8"))

for row in worksheet.iter_rows(min_row=2):
    for col in range(1, worksheet.max_column):
        out_files[col - 1].write("{}={}\n".format(row[0].value, row[col].value.replace("\n", "\\n\\\n  ")))
